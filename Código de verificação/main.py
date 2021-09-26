'''----------- IMPORTS -----------'''

import openpyxl
import sys

'''----------- IMPORTS -----------'''

'''----------- VARIABLES -----------'''

DataList = []
index = 0
MainList = []
idade = []
questoes = []

wb = openpyxl.load_workbook(filename=r"C:\Users\Inácio Rodrigues\Desktop\Nova pasta\base_dados_simulada.xlsx")

'''----------- VARIABLES -----------'''

'''----------- CODE -----------'''

wb.sheetnames

print('\n--- Dados retirados do Excel --- \n')

for row in wb['Worksheet'].iter_rows(values_only = True):
	index += 1
	try:
		DataList = list(row)
	
		MainList.append(DataList)
	except:
		pass

for i in range(len(MainList)):
	for a in range(len(MainList[i])):
		if(a == 20 and i > 0):
			idade.append(int(MainList[i][a]))

media = round(sum(idade)/len(idade),0)

print('Média de idades é igual a ' + str(int(media)) + ' anos.\n')

for i in range(len(MainList)):
	for a in range(len(MainList[i])):
		if(a > 20 and a < 197 and i == 0):
			questoes.append([MainList[i][a]])

		if(a > 20 and a < 197 and i > 0):
			value = a - 21
			add = True
			for h in range(len(questoes[value])):
				try:
					if(str(questoes[value][h]) == str(MainList[i][a])):
						add = False
						questoes[value][h+1] = questoes[value][h+1] + 1
				except:
					pass
			if(add):
				questoes[value].append(str(MainList[i][a]))
				questoes[value].append(1)

for i in range(len(questoes)):
	soma = 0
	for a in range(len(questoes[i])+100):
		try:
			soma = soma + questoes[i][a]
		except:
			pass

	for a in range(len(questoes[i])+100):
		try:
			questoes[i].insert(a+1,(str(round(((questoes[i][a]/soma)*100),2))+'%'))
		except:
			pass

for i in range(len(questoes)):
	print('\nRespostas para questão ' + str(questoes[i][0]) + ':')
	
	for a in range(1, len(questoes[i])+1,3):
		try:
			print('Numero de respostas para alternativa -' + str(questoes[i][a]) + '- foram: ' + str(questoes[i][a+1]) + ' com porcentagem de ' + str(questoes[i][a+2]))
		except:
			pass